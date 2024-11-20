import pdfplumber
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from copy import deepcopy
import re
import tkinter as tk
from tkinter import filedialog, messagebox
import os

# Função para extrair o contrato, OS, data e Num do texto
# Função para extrair o contrato, OS, data e Num do texto
# Função atualizada para extrair contrato, OS, data e Num do texto
def extract_contract_os_num(text):
    # Ajustar regex para aceitar contratos com zeros extras no início
    contract_match = re.search(r'Contrato: ([\w\-]+ \d+\/\d{2})', text)
    # Ajustar regex para aceitar OS com letras e números
    os_match = re.search(r'OS: (\w+)', text)
    # Mantém a regex para a data e número do relatório
    date_match = re.search(r'Data: (\d{2}\/\d{2}\/\d{4})', text)
    num_match = re.search(r'Num: (\d{3}\/\d{4})', text)

    contract = contract_match.group(1) if contract_match else None
    os_value = os_match.group(1) if os_match else None
    date = date_match.group(1) if date_match else None
    num_value = num_match.group(1) if num_match else None

    # Adicionando prints de debug para verificar os valores extraídos
    print(f"Contrato: {contract}")
    print(f"OS: {os_value}")
    print(f"Data: {date}")
    print(f"Num: {num_value}")

    return contract, os_value, date, num_value

# Função atualizada para extrair a disciplina
def extract_discipline(text):
    # Ajustar o regex para capturar qualquer texto entre "disciplina de" e ":"
    discipline_match = re.search(r"disciplina de (.+?):", text, re.IGNORECASE)
    discipline = discipline_match.group(1).strip().upper() if discipline_match else None

    # Log de debug para a disciplina
    print(f"Disciplina encontrada: {discipline}")

    return discipline





# Função para extrair NCs do PDF
def extract_ncs_from_pdf(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        text = ""
        for page in pdf.pages:
            text += page.extract_text()

    lines = text.split("\n")
    ncs = []
    current_nc = []
    collecting = False

    end_keywords = ["Relatório", "Cliente", "Folha", "Concessionária", "Num:", "Etapa", "OM", "Oportunidade de Melhorias", "FOR-713"]

    for line in lines:
        line = line.strip()

        if line.startswith(tuple(f"{i}." for i in range(1, 100))) and "Requisito não atendido:" in line:
            if collecting:
                ncs.append("\n".join(current_nc))
            collecting = True
            current_nc = [line]

        elif collecting and current_nc and not any(keyword in line for keyword in end_keywords):
            current_nc.append(line)

        elif collecting and any(keyword in line for keyword in end_keywords):
            if current_nc:
                ncs.append("\n".join(current_nc))
                current_nc = []
            collecting = False

    if current_nc:
        ncs.append("\n".join(current_nc))

    return ncs, text

# Função para copiar imagens entre abas
def copy_images(source_sheet, target_sheet):
    if source_sheet._images:
        for img in source_sheet._images:
            new_img = deepcopy(img)
            target_sheet.add_image(new_img)

def extract_inspection_text(text):
    # Procurar um padrão que comece com "Foram inspecionados" e continue até "listas de verificação."
    inspection_match = re.search(r"(Foram inspecionados[\s\S]+?listas de verificação\.)", text)
    if inspection_match:
        return inspection_match.group(0)
    return None



def extract_for_number(text):
    # Procurar o número do FOR no formato "FOR-XXX"
    for_match = re.search(r"FOR-(\d+)", text)
    if for_match:
        return f"FOR-{for_match.group(1)}"
    return None


# Função para salvar NCs no template de Excel e preencher a aba DISCIPLINA
def save_ncs_to_template_excel(ncs, template_path, excel_path, contract, os_value, date, num_value, inspection_text, discipline):
    wb = load_workbook(template_path)
    template_sheet = wb.active

    disciplina_sheet = wb['DISCIPLINA']  # Selecionar a aba DISCIPLINA

    for i, nc in enumerate(ncs, start=1):
        new_sheet = wb.copy_worksheet(template_sheet)
        new_sheet.title = f"{i}"

        copy_images(template_sheet, new_sheet)

        # Adicionar a NC na célula A14
        new_sheet.cell(row=14, column=1, value=nc)
        
        #essa aqui é pra colocar o numero do relatorio nas NCs
        new_sheet.cell(row=7, column=16, value=num_value)
        
        new_sheet.cell(row=10, column=1, value=discipline)

        # Inserir o contrato e OS na célula A8 da aba NC
        new_sheet.cell(row=8, column=1, value=f"{contract} - {os_value}")

        # Inserir o número da NC na célula A12
        new_sheet.cell(row=12, column=1, value=f"{i}")

        # Inserir a data na célula P9 da aba NC
        new_sheet.cell(row=9, column=16, value=date)

        # Preencher a aba DISCIPLINA com informações da aba NC
        disciplina_sheet.cell(row=7, column=1, value=f"{contract} - {os_value}")  # A8 da NC vai para A7 da DISCIPLINA
        disciplina_sheet.cell(row=6, column=16, value=num_value)  # P7 da NC vai para P6 da DISCIPLINA
        disciplina_sheet.cell(row=8, column=16, value=date)  # P9 da NC vai para P8 da DISCIPLINA

        # Adicionar o texto de inspeção na célula A11 da aba DISCIPLINA
        # if inspection_text:
        #     disciplina_sheet.cell(row=11, column=1, value=inspection_text)
        
        # Adicionar a disciplina extraída na célula A9 da aba DISCIPLINA
        if discipline:
            disciplina_sheet.cell(row=9, column=1, value=discipline)

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    wb.save(excel_path)



# Funções da interface gráfica
def select_pdf():
    file_path = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
    if file_path:
        pdf_path_var.set(file_path)

def select_output_folder():
    folder_path = filedialog.askdirectory()
    if folder_path:
        output_folder_var.set(folder_path)

def select_template():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        template_path_var.set(file_path)


def process_pdf():
    pdf_path = pdf_path_var.get()
    output_folder = output_folder_var.get()
    template_path = template_path_var.get()

    if not pdf_path or not output_folder or not template_path:
        messagebox.showwarning("Erro", "Por favor, selecione um arquivo PDF, um modelo de Excel e uma pasta de saída.")
        return
    
    try:
        ncs, pdf_text = extract_ncs_from_pdf(pdf_path)
        contract, os_value, date, num_value = extract_contract_os_num(pdf_text)
        inspection_text = extract_inspection_text(pdf_text)
        discipline = extract_discipline(pdf_text)  # Extraindo a disciplina

        if not contract or not os_value or not date or not discipline:
            messagebox.showwarning("Erro", "Contrato, OS, data ou disciplina não foram encontrados no PDF.")
            return

        # Usar o nome do arquivo PDF como base para o nome do Excel
        pdf_filename = os.path.splitext(os.path.basename(pdf_path))[0]  # Obtém o nome do arquivo PDF sem extensão

        # Adicionar a disciplina (em maiúsculas) ao nome do arquivo
        output_excel_filename = f"{pdf_filename}_{discipline.upper()}.xlsx"

        if ncs:
            output_excel_path = os.path.join(output_folder, output_excel_filename)  # Nome do Excel com disciplina
            save_ncs_to_template_excel(ncs, template_path, output_excel_path, contract, os_value, date, num_value, inspection_text, discipline)
            messagebox.showinfo("Sucesso", f"Arquivo Excel '{output_excel_filename}' criado com {len(ncs)} NCs, cada uma em uma aba.")
        else:
            messagebox.showwarning("Nenhuma NC", "Nenhuma NC foi encontrada no arquivo PDF.")
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro: {str(e)}")




# Configuração da interface gráfica
app = tk.Tk()
app.title("Processador de NCs")
app.geometry("500x400")

pdf_path_var = tk.StringVar()
output_folder_var = tk.StringVar()
template_path_var = tk.StringVar()

tk.Label(app, text="Selecione o PDF com as NCs:").pack(pady=10)
tk.Entry(app, textvariable=pdf_path_var, width=50, state="readonly").pack(pady=5)
tk.Button(app, text="Selecionar PDF", command=select_pdf).pack(pady=5)

tk.Label(app, text="Selecione a pasta de saída:").pack(pady=10)
tk.Entry(app, textvariable=output_folder_var, width=50, state="readonly").pack(pady=5)
tk.Button(app, text="Selecionar Pasta", command=select_output_folder).pack(pady=5)

tk.Label(app, text="Selecione o modelo de Excel:").pack(pady=10)
tk.Entry(app, textvariable=template_path_var, width=50, state="readonly").pack(pady=5)
tk.Button(app, text="Selecionar Modelo", command=select_template).pack(pady=5)

tk.Button(app, text="Iniciar Processamento", command=process_pdf, bg="green", fg="white").pack(pady=20)

app.mainloop()
