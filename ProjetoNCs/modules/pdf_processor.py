import tkinter as tk
from tkinter import filedialog, messagebox
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from copy import deepcopy
import re
import pdfplumber


class PDFProcessor:
    def __init__(self, root):
        self.root = root
        self.root.title("Processador de NCs")
        self.root.geometry("500x400")

        self.output_folder_var = tk.StringVar()
        self.template_path_var = tk.StringVar()
        self.pdf_paths_var = tk.StringVar()

        tk.Label(root, text="Selecione os PDFs:").pack(pady=10)
        tk.Entry(root, textvariable=self.pdf_paths_var, width=50, state="readonly").pack(pady=5)
        tk.Button(root, text="Selecionar PDFs", command=self.select_pdfs).pack(pady=5)

        tk.Label(root, text="Selecione a pasta de saída:").pack(pady=10)
        tk.Entry(root, textvariable=self.output_folder_var, width=50, state="readonly").pack(pady=5)
        tk.Button(root, text="Selecionar Pasta", command=self.select_output_folder).pack(pady=5)

        tk.Label(root, text="Selecione o modelo de Excel:").pack(pady=10)
        tk.Entry(root, textvariable=self.template_path_var, width=50, state="readonly").pack(pady=5)
        tk.Button(root, text="Selecionar Modelo", command=self.select_template).pack(pady=5)

        tk.Button(root, text="Processar PDFs", command=self.process_pdfs, bg="green", fg="white").pack(pady=20)
        
    

    def select_output_folder(self):
        folder_path = filedialog.askdirectory()
        if folder_path:
            self.output_folder_var.set(folder_path)

    def select_template(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            self.template_path_var.set(file_path)

    def select_pdfs(self):
        file_paths = filedialog.askopenfilenames(filetypes=[("PDF files", "*.pdf")])
        if file_paths:
            self.pdf_paths_var.set("; ".join(file_paths))

    def process_pdfs(self):
        pdf_paths = self.pdf_paths_var.get().split("; ")
        output_folder = self.output_folder_var.get()
        template_path = self.template_path_var.get()

        if not pdf_paths or not output_folder or not template_path:
            messagebox.showwarning("Erro", "Por favor, selecione arquivos PDF, um modelo de Excel e uma pasta de saída.")
            return

        for pdf_path in pdf_paths:
            try:
                ncs, pdf_text = extract_ncs_from_pdf(pdf_path)
                contract, os_value, date, num_value, rev_value = extract_contract_os_num(pdf_text)
                discipline = extract_discipline(pdf_text)
                codes = extract_document_codes(pdf_text)

                if not contract or not os_value or not date or not discipline:
                    messagebox.showwarning("Erro", f"Contrato, OS, data ou disciplina não foram encontrados no PDF '{os.path.basename(pdf_path)}'.")
                    continue
                if rev_value is None:
                    rev_value = "00"  # Valor padrão

                split_num_relatorio = num_value.split("/")
                num_relatorio = split_num_relatorio[0]
                ano_relatorio = split_num_relatorio[1]

                print(f"Contract: {contract}, OS: {os_value}, Date: {date}, Num: {num_value}, Rev: {rev_value}")

                pdf_filename = os.path.splitext(os.path.basename(pdf_path))[0]

                output_excel_filename = f"FOR875 RI{num_relatorio}-{ano_relatorio}-R{rev_value}_{discipline.upper()}.xlsx"
                output_excel_path = os.path.join(output_folder, output_excel_filename)

                if ncs:
                    save_ncs_to_template_excel(ncs, template_path, output_excel_path, contract, os_value, date, num_value, discipline)

                    wb = load_workbook(output_excel_path)
                    save_codes_to_ldd_sheet(codes, wb)
                    wb.save(output_excel_path)

                    print(f"Arquivo Excel '{output_excel_filename}' criado com sucesso.")
                else:
                    print(f"Nenhuma NC foi encontrada no arquivo PDF '{pdf_filename}'.")

            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao processar o PDF '{os.path.basename(pdf_path)}': {str(e)}")
                return  # Parar o processamento caso ocorra um erro

        messagebox.showinfo("Sucesso", "Todos os PDFs foram processados com sucesso!")



# Função para extrair o contrato, OS, data e Num do texto
def extract_contract_os_num(text):
    contract_match = re.search(r'Contrato: ([A-Z\s-]+ \d+\/\d{2})', text)
    os_match = re.search(r'OS: (\w+)', text)
    date_match = re.search(r'Data: (\d{2}\/\d{2}\/\d{4})', text)
    num_match = re.search(r'Num: (\d{3}\/\d{4})', text)
    rev_match = re.search(r'Rev\.\s*:\s*(\d+)', text)  # Captura o valor após "Rev.:"

    contract = contract_match.group(1) if contract_match else None
    os_value = os_match.group(1) if os_match else None
    date = date_match.group(1) if date_match else None
    num_value = num_match.group(1) if num_match else None
    rev_value = rev_match.group(1) if rev_match else None  # Captura o valor de Rev.

    return contract, os_value, date, num_value, rev_value

def extract_discipline(text):
    discipline_match = re.search(r"disciplina de (.+?):", text, re.IGNORECASE)
    discipline = discipline_match.group(1).strip().upper() if discipline_match else None
    return discipline

def extract_document_codes(text):
    clean_text = re.sub(r'\s+', ' ', text)

    patterns = [
        r'\bERM-\d{3}[A-Z]{2}-\d{3}\+\d{3}-[A-Z]{3}-[A-Z]{3}-[A-Z]{2}-[A-Z0-9]{2}-\d{3}_R\d{2,3}[A-Z]?\b',
        r'\bECA-\d{3}[A-Z]{2}-\d{3}-\d{3}-[A-Z]{3}-[A-Z]{3}-[A-Z]{2}-[A-Z0-9]{2}-\d{3}-R\d{2,3}[A-Za-z]?\b',
        r'\bERM-\d{3}[A-Z]{2}-\d{3}\-\d{3}-[A-Z]{3}-[A-Z]{3}-[A-Z]{2}-[A-Z0-9]{2}-\d{3}_R\d{2,3}[A-Z]?\b',
        r'\bERM-\d{3}[A-Z]{2}-\d{3}-\d{3}-[A-Z]{3}-[A-Z]{3}-[A-Z]{2}-[A-Z0-9]{2}-\d{3}-R\d{2,3}[A-Z]?\b'
    ]

    total_docs_match = re.search(r'Total\s*=\s*(\d+)', text)
    total_docs = int(total_docs_match.group(1)) if total_docs_match else 7

    codes_set = set()

    for pattern in patterns:
        found_codes = re.findall(pattern, clean_text)
        for code in found_codes:
            if len(codes_set) < total_docs:
                codes_set.add(code)
            if len(codes_set) == total_docs:
                break
        if len(codes_set) == total_docs:
            break

    codes = list(codes_set)
    return codes

def extract_ncs_from_pdf(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        text = ""
        for page in pdf.pages:
            text += page.extract_text()
    
    lines = text.split("\n")
    ncs = []
    current_nc = []
    collecting = False
    inside_header = False  # Controle de cabeçalho para remoção

    # Expressão regular para identificar o início de cada NC
    nc_start_pattern = re.compile(r'^\d+\.\s+Requisito não atendido')

    # Padrões para detectar linhas específicas do cabeçalho
    header_patterns = [
        r"FOR-713 – Relatório de Inspeção", 
        r"IAT – Instituto Alphageos de Tecnologia",
        r"Cliente:", r"Contrato:", r"Folha:", r"Rev.:", r"Data:", r"OS:", r"Etapa:", r"Num:", 
        r"ECOVIAS DO ARAGUAIA", r"S\.A", r"ao km \d+\+\d+"
    ]

    for line in lines:
        line = line.strip()

        # Ativa a remoção de cabeçalho se a linha corresponde a um padrão de cabeçalho
        if any(re.search(pattern, line) for pattern in header_patterns):
            inside_header = True
            continue  # Pula linhas que pertencem ao cabeçalho
        
        # Desativa o cabeçalho ao encontrar uma linha que não pertence ao cabeçalho
        if inside_header and not any(re.search(pattern, line) for pattern in header_patterns):
            inside_header = False  # Fim do cabeçalho

        # Ignora linhas dentro do cabeçalho
        if inside_header:
            continue

        # Detecta o início de uma nova NC
        if nc_start_pattern.match(line):
            if collecting:
                # Salva a NC completa na lista antes de iniciar uma nova
                ncs.append("\n".join(current_nc))
            # Inicia uma nova NC
            collecting = True
            current_nc = [line]
        elif collecting:
            # Continua a coletar linhas para o NC atual
            current_nc.append(line)

    # Adiciona a última NC capturada, se existir
    if current_nc:
        ncs.append("\n".join(current_nc))

    return ncs, text



def copy_images(source_sheet, target_sheet):
    if source_sheet._images:
        for img in source_sheet._images:
            new_img = deepcopy(img)
            target_sheet.add_image(new_img)

def save_ncs_to_template_excel(ncs, template_path, excel_path, contract, os_value, date, num_value, discipline):
    wb = load_workbook(template_path)
    template_sheet = wb.active
    disciplina_sheet = wb['DISCIPLINA']

    for i, nc in enumerate(ncs, start=1):
        new_sheet = wb.copy_worksheet(template_sheet)
        new_sheet.title = f"NC {i}"
        copy_images(template_sheet, new_sheet)

        new_sheet.cell(row=14, column=1, value=nc)
        new_sheet.cell(row=7, column=16, value=num_value)
        new_sheet.cell(row=10, column=1, value=discipline)
        new_sheet.cell(row=7, column=1, value=f"{contract} - {os_value}")
        new_sheet.cell(row=9, column=16, value=date)
        new_sheet.cell(row=6, column=16, value=discipline)
        new_sheet.cell(row=12, column=1, value=f"{i}")
        disciplina_sheet['A9'] = discipline

    wb.save(excel_path)

def save_codes_to_ldd_sheet(codes, wb):
    ldd_sheet = wb['LDD']

    for i, code in enumerate(codes, start=2):
        ldd_sheet.cell(row=i, column=1, value=code)


if __name__ == "__main__":
    root = tk.Tk()
    app = PDFProcessor(root)
    root.mainloop()
