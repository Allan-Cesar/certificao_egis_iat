import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import os

class ResponseExtractorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Extrair Respostas")
        self.root.geometry("400x400")
        
        # Campo de entrada para os arquivos de entrada
        tk.Label(root, text="Arquivos de Entrada (.xlsx)").pack(pady=10)
        self.input_entry = tk.Entry(root, width=40)
        self.input_entry.pack(pady=5)
        tk.Button(root, text="Selecionar", command=self.select_input_files).pack(pady=5)
        
        # Campo para o arquivo modelo
        tk.Label(root, text="Arquivo Modelo (.xlsx)").pack(pady=10)
        self.model_entry = tk.Entry(root, width=40)
        self.model_entry.pack(pady=5)
        tk.Button(root, text="Selecionar", command=self.select_model_file).pack(pady=5)
        
        # Campo de saída para a pasta de saída
        tk.Label(root, text="Pasta de Saída").pack(pady=10)
        self.output_entry = tk.Entry(root, width=40)
        self.output_entry.pack(pady=5)
        tk.Button(root, text="Selecionar Pasta", command=self.select_output_folder).pack(pady=5)
        
        # Botão para extrair respostas
        tk.Button(root, text="Extrair Respostas", command=self.run_extraction).pack(pady=20)
    
    def select_input_files(self):
        input_files = filedialog.askopenfilenames(
            title="Selecione os Arquivos de Entrada",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if input_files:
            self.input_entry.delete(0, tk.END)
            self.input_entry.insert(0, ";".join(input_files))
    
    def select_model_file(self):
        model_file = filedialog.askopenfilename(
            title="Selecione o Arquivo Modelo",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if model_file:
            self.model_entry.delete(0, tk.END)
            self.model_entry.insert(0, model_file)

    def select_output_folder(self):
        output_folder = filedialog.askdirectory(
            title="Selecione a Pasta de Saída"
        )
        if output_folder:
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, output_folder)
    
    def run_extraction(self):
        input_files = self.input_entry.get().split(";")
        model_file = self.model_entry.get()
        output_folder = self.output_entry.get()
        
        if not input_files or not model_file or not output_folder:
            messagebox.showerror("Erro", "Por favor, selecione os arquivos de entrada, o arquivo modelo e a pasta de saída.")
            return
        
        output_file = os.path.join(output_folder, "Respostas_Consolidadas.xlsx")
        
        try:
            consolidate_responses_data(input_files, model_file, output_file)
            messagebox.showinfo("Sucesso", f"Dados extraídos com sucesso para {output_file}")
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro ao extrair os dados: {e}")

def consolidate_responses_data(input_files, model_file, output_file):
    # Abre o arquivo modelo
    wb_output = openpyxl.load_workbook(model_file)
    ws_output = wb_output.active  # Assume que os dados serão preenchidos na primeira aba
    
    start_row = 4  # Linha inicial para inserção dos dados
    current_row = start_row
    
    # Itera sobre cada arquivo de entrada
    for input_file in input_files:
        # Abre a planilha de entrada
        wb_input = openpyxl.load_workbook(input_file)
        
        # Itera sobre as abas NC (NC 1, NC 2, ...)
        for sheet_name in wb_input.sheetnames:
            if sheet_name.startswith("NC "):
                ws = wb_input[sheet_name]
                
                # Extrai os valores das células desejadas
                nc_num = ws["P7"].value if ws["P7"].value else ""
                disciplina = ws["A10"].value if ws["A10"].value else ""
                nc_text = ws["A14"].value if ws["A14"].value else ""
                resposta_1 = ws["A16"].value if ws["A16"].value else ""
                resposta_2 = ws["A18"].value if ws["A18"].value else ""
                
                # Adiciona os dados nas colunas A, D, L, M e N da linha atual
                ws_output[f"A{current_row}"] = nc_num
                ws_output[f"D{current_row}"] = disciplina
                ws_output[f"L{current_row}"] = nc_text
                ws_output[f"R{current_row}"] = resposta_1
                ws_output[f"S{current_row}"] = resposta_2
                
                # Move para a próxima linha para a próxima entrada
                current_row += 1

    # Salva o arquivo de saída consolidado
    wb_output.save(output_file)

