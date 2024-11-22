import pdfplumber
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from copy import deepcopy
import re
import tkinter as tk
from tkinter import filedialog, messagebox
import os

# Função para extrair o contrato, OS, data e Num do texto
def extract_contract_os_num(text):
    contract_match = re.search(r'Contrato: ([A-Z\s-]+ \d+\/\d{2})', text)
    os_match = re.search(r'OS: (\w+)', text)
    date_match = re.search(r'Data: (\d{2}\/\d{2}\/\d{4})', text)
    num_match = re.search(r'Num: (\d{3}\/\d{4})', text)

    contract = contract_match.group(1) if contract_match else None
    os_value = os_match.group(1) if os_match else None
    date = date_match.group(1) if date_match else None
    num_value = num_match.group(1) if num_match else None

    return contract, os_value, date, num_value

def extract_discipline(text):
    discipline_match = re.search(r"disciplina de (.+?):", text, re.IGNORECASE)
    discipline = discipline_match.group(1).strip().upper() if discipline_match else None
    return discipline

def extract_document_codes(text):
    # Substituir múltiplos espaços ou quebras de linha por um único espaço para manter a estrutura intacta
    clean_text = re.sub(r'\s+', ' ', text)

    # Expressões regulares que cobrem os diferentes formatos de código
    patterns = [
        r'\bERM-\d{3}[A-Z]{2}-\d{3}\+\d{3}-[A-Z]{3}-[A-Z]{3}-[A-Z]{2}-[A-Z0-9]{2}-\d{3}_R\d{2,3}[A-Z]?\b',  # Padrão ERM com "+"
        r'\bECA-\d{3}[A-Z]{2}-\d{3}-\d{3}-[A-Z]{3}-[A-Z]{3}-[A-Z]{2}-[A-Z0-9]{2}-\d{3}-R\d{2,3}[A-Za-z]?\b',  # Padrão ECA
        r'\bERM-\d{3}[A-Z]{2}-\d{3}\-\d{3}-[A-Z]{3}-[A-Z]{3}-[A-Z]{2}-[A-Z0-9]{2}-\d{3}_R\d{2,3}[A-Z]?\b',  # Padrão ERM de projeto estrutural e memorial de cálculo
        r'\bERM-\d{3}[A-Z]{2}-\d{3}-\d{3}-[A-Z]{3}-[A-Z]{3}-[A-Z]{2}-[A-Z0-9]{2}-\d{3}-R\d{2,3}[A-Z]?\b'
        
    ]
    
    # Encontrar o número de documentos na lista (por exemplo, "Total = 7")
    total_docs_match = re.search(r'Total\s*=\s*(\d+)', text)
    total_docs = int(total_docs_match.group(1)) if total_docs_match else 7  # Se não encontrar, assume 7 por padrão

    # Para armazenar os códigos encontrados (remover duplicatas com conjunto)
    codes_set = set()
    
    # Aplicar cada padrão de regex para capturar os códigos
    for pattern in patterns:
        found_codes = re.findall(pattern, clean_text)
        for code in found_codes:
            if len(codes_set) < total_docs:  # Adiciona até o limite definido por "Total"
                codes_set.add(code)
            if len(codes_set) == total_docs:  # Se atingir o limite, sai do loop
                break
        if len(codes_set) == total_docs:
            break

    codes = list(codes_set)

    # Log para depuração
    print(f"Total de códigos encontrados: {len(codes)}")
    print(f"Códigos encontrados: {codes}")

    return codes


def extract_ncs_from_pdf(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        text = ""
        for page in pdf.pages:
            text += page.extract_text()

    lines = text.split("\n")
    ncs = []
    current_nc = []
    collecting = False

    end_keywords = ["Relatório", "Cliente", "Folha", "Concessionária", "Num:", "Etapa", "OM", "Oportunidade de Melhorias", "FOR-713"]

    for line in lines:
        line = line.strip()

        if line.startswith(tuple(f"{i}." for i in range(1, 100))) and "Requisito não atendido:" in line:
            if collecting:
                ncs.append("\n".join(current_nc))
            collecting = True
            current_nc = [line]

        elif collecting and current_nc and not any(keyword in line for keyword in end_keywords):
            current_nc.append(line)

        elif collecting and any(keyword in line for keyword in end_keywords):
            if current_nc:
                ncs.append("\n".join(current_nc))
                current_nc = []
            collecting = False

    if current_nc:
        ncs.append("\n".join(current_nc))

    return ncs, text

def copy_images(source_sheet, target_sheet):
    if source_sheet._images:
        for img in source_sheet._images:
            new_img = deepcopy(img)
            target_sheet.add_image(new_img)

def save_ncs_to_template_excel(ncs, template_path, excel_path, contract, os_value, date, num_value, inspection_text, discipline):
    wb = load_workbook(template_path)
    template_sheet = wb.active
    disciplina_sheet = wb['DISCIPLINA']

    for i, nc in enumerate(ncs, start=1):
        new_sheet = wb.copy_worksheet(template_sheet)
        new_sheet.title = f"NC {i}"
        copy_images(template_sheet, new_sheet)

        new_sheet.cell(row=14, column=1, value=nc)
        new_sheet.cell(row=7, column=16, value=num_value)
        new_sheet.cell(row=10, column=1, value=discipline)
        new_sheet.cell(row=8, column=1, value=f"{contract} - {os_value}")
        new_sheet.cell(row=12, column=1, value=f"{i}")
        new_sheet.cell(row=9, column=16, value=date)

        disciplina_sheet.cell(row=7, column=1, value=f"{contract} - {os_value}")
        disciplina_sheet.cell(row=6, column=16, value=num_value)
        disciplina_sheet.cell(row=8, column=16, value=date)

        if discipline:
            disciplina_sheet.cell(row=9, column=1, value=discipline)

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    wb.save(excel_path)

def save_codes_to_ldd_sheet(codes, workbook):
    ldd_sheet = workbook['LDD']
    for i, code in enumerate(codes, start=2):
        ldd_sheet.cell(row=i, column=1, value=code)
        print(f"Salvando código {code} na linha {i}")

def process_pdfs():
    # Usa os arquivos já selecionados e mostrados no campo `pdf_paths_var`
    pdf_paths = pdf_paths_var.get().split("; ")  
    output_folder = output_folder_var.get()
    template_path = template_path_var.get()

    if not pdf_paths or not output_folder or not template_path:
        messagebox.showwarning("Erro", "Por favor, selecione arquivos PDF, um modelo de Excel e uma pasta de saída.")
        return

    for pdf_path in pdf_paths:
        try:
            ncs, pdf_text = extract_ncs_from_pdf(pdf_path)
            contract, os_value, date, num_value = extract_contract_os_num(pdf_text)
            inspection_text = extract_inspection_text(pdf_text)
            discipline = extract_discipline(pdf_text)  
            codes = extract_document_codes(pdf_text)

            if not contract or not os_value or not date or not discipline:
                messagebox.showwarning("Erro", f"Contrato, OS, data ou disciplina não foram encontrados no PDF '{os.path.basename(pdf_path)}'.")
                continue

            pdf_filename = os.path.splitext(os.path.basename(pdf_path))[0]  
            output_excel_filename = f"{pdf_filename}_{discipline.upper()}.xlsx"
            output_excel_path = os.path.join(output_folder, output_excel_filename)

            if ncs:
                save_ncs_to_template_excel(ncs, template_path, output_excel_path, contract, os_value, date, num_value, inspection_text, discipline)
                
                wb = load_workbook(output_excel_path)
                save_codes_to_ldd_sheet(codes, wb)
                wb.save(output_excel_path)

                print(f"Arquivo Excel '{output_excel_filename}' criado com sucesso.")
            else:
                print(f"Nenhuma NC foi encontrada no arquivo PDF '{pdf_filename}'.")

        except Exception as e:
            print(f"Erro ao processar o PDF '{os.path.basename(pdf_path)}': {str(e)}")

# Função para selecionar a pasta de saída
def select_output_folder():
    folder_path = filedialog.askdirectory()  # Abre o seletor de pastas
    if folder_path:
        output_folder_var.set(folder_path)  # Define o caminho da pasta escolhida

# Função para selecionar o modelo de Excel
def select_template():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        template_path_var.set(file_path)
        
def extract_inspection_text(text):
    # Procurar um padrão que comece com "Foram inspecionados" e continue até "listas de verificação."
    inspection_match = re.search(r"(Foram inspecionados[\s\S]+?listas de verificação\.)", text)
    if inspection_match:
        return inspection_match.group(0)
    return None

# Função para selecionar múltiplos PDFs
def select_pdfs():
    file_paths = filedialog.askopenfilenames(filetypes=[("PDF files", "*.pdf")])
    if file_paths:
        pdf_paths_var.set("; ".join(file_paths))  # Mostra os arquivos selecionados

# Interface gráfica
app = tk.Tk()
app.title("Processador de NCs")
app.geometry("500x400")

output_folder_var = tk.StringVar()
template_path_var = tk.StringVar()
pdf_paths_var = tk.StringVar()  # Variável para armazenar os PDFs selecionados

tk.Label(app, text="Selecione os PDFs:").pack(pady=10)
tk.Entry(app, textvariable=pdf_paths_var, width=50, state="readonly").pack(pady=5)
tk.Button(app, text="Selecionar PDFs", command=select_pdfs).pack(pady=5)

tk.Label(app, text="Selecione a pasta de saída:").pack(pady=10)
tk.Entry(app, textvariable=output_folder_var, width=50, state="readonly").pack(pady=5)
tk.Button(app, text="Selecionar Pasta", command=select_output_folder).pack(pady=5)

tk.Label(app, text="Selecione o modelo de Excel:").pack(pady=10)
tk.Entry(app, textvariable=template_path_var, width=50, state="readonly").pack(pady=5)
tk.Button(app, text="Selecionar Modelo", command=select_template).pack(pady=5)

tk.Button(app, text="Processar PDFs", command=process_pdfs, bg="green", fg="white").pack(pady=20)

app.mainloop()
